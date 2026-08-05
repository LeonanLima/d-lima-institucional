# -*- coding: utf-8 -*-
"""Gera a planilha de flecha NBR 6118 (dashboard + viga + laje macica + trelicada)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

AZUL = "1A3C6E"
AMARELO = "FFF2CC"
CINZA = "F2F2F2"
VERDE_F = "C6EFCE"; VERDE_T = "006100"
VERM_F = "FFC7CE"; VERM_T = "9C0006"
borda = Border(*[Side(style="thin", color="B0B0B0")] * 4)

F_TIT = Font(bold=True, size=14, color="FFFFFF")
F_SEC = Font(bold=True, size=11, color="FFFFFF")
F_IN = Font(color="0000FF")
F_CALC = Font(color="000000")
F_LINK = Font(color="008000")
F_NOTA = Font(size=8, italic=True, color="808080")


def titulo(ws, texto, sub):
    ws.merge_cells("A1:E1"); ws.merge_cells("A2:E2")
    ws["A1"] = texto; ws["A1"].font = F_TIT
    ws["A1"].fill = PatternFill("solid", start_color=AZUL)
    ws["A2"] = sub; ws["A2"].font = F_NOTA
    for col, w in zip("ABCDE", (42, 14, 10, 52, 12)):
        ws.column_dimensions[col].width = w


def secao(ws, r, texto):
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value=texto)
    c.font = F_SEC; c.fill = PatternFill("solid", start_color=AZUL)
    return r + 1


def linha(ws, r, rotulo, valor, unid="", nota="", tipo="calc", fmt="0.000"):
    ws.cell(row=r, column=1, value=rotulo).border = borda
    c = ws.cell(row=r, column=2, value=valor)
    c.border = borda
    if isinstance(valor, (int, float)) or (isinstance(valor, str) and valor.startswith("=")):
        c.number_format = fmt
    if tipo == "in":
        c.font = F_IN; c.fill = PatternFill("solid", start_color=AMARELO)
        c.protection = Protection(locked=False)
    else:
        c.font = F_CALC; c.fill = PatternFill("solid", start_color=CINZA)
    ws.cell(row=r, column=3, value=unid).font = F_NOTA
    ws.cell(row=r, column=4, value=nota).font = F_NOTA
    return r + 1


def cond_status(ws, faixa):
    ws.conditional_formatting.add(faixa, CellIsRule(
        operator="equal", formula=['"OK"'],
        fill=PatternFill("solid", start_color=VERDE_F), font=Font(bold=True, color=VERDE_T)))
    ws.conditional_formatting.add(faixa, CellIsRule(
        operator="equal", formula=['"NAO PASSA"'],
        fill=PatternFill("solid", start_color=VERM_F), font=Font(bold=True, color=VERM_T)))


def dv_lista(ws, cel, opcoes):
    dv = DataValidation(type="list", formula1=f'"{opcoes}"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(cel)


VINC = "Biapoiada,Apoiada-Engastada,Biengastada,Balanco"
AGREG = "Basalto/Diabasio,Granito/Gnaisse,Calcario,Arenito"
F_K = '=IF({v}="Biapoiada",5/384,IF({v}="Apoiada-Engastada",1/185,IF({v}="Biengastada",1/384,1/8)))'
F_CM = '=IF({v}="Biapoiada",1/8,IF({v}="Apoiada-Engastada",9/128,IF({v}="Biengastada",1/24,1/2)))'
F_AE = ('=IF({a}="Basalto/Diabasio",1.2,IF({a}="Granito/Gnaisse",1,'
        'IF({a}="Calcario",0.9,0.7)))')


def bloco_flechas_verifs(ws, r, R, celvao_cm, cel_aimed, cel_qlin, cel_plin, cel_psi2,
                         cel_t0, cel_rho, cel_parede, cel_vinc):
    """Escreve flechas diferida/total + verificacoes. Devolve (r, refs p/ dashboard)."""
    r = secao(ws, r, "FLECHA DIFERIDA E TOTAL (NBR 6118, 17.3.2.1.2)")
    R["xi0"] = r; r = linha(ws, r, "Coef. xi(t0)", f"=IF({cel_t0}>=70,2,0.68*(0.996^{cel_t0})*({cel_t0}^0.32))", "-", "xi(inf)=2")
    R["af"] = r; r = linha(ws, r, "alfa_f = (2 - xi0)/(1 + 50*rho')", f"=(2-B{R['xi0']})/(1+50*{cel_rho})", "-")
    R["adif"] = r; r = linha(ws, r, "Flecha diferida adicional", f"={cel_aimed}*B{R['af']}", "cm")
    R["atot"] = r; r = linha(ws, r, "FLECHA TOTAL (imediata + diferida)", f"={cel_aimed}*(1+B{R['af']})", "cm")
    R["cx"] = r; r = linha(ws, r, "Relacao vao/flecha (C/x)", f"=IF(B{R['atot']}=0,0,{celvao_cm}/B{R['atot']})", "-", "compare com o software", fmt="0")
    r += 1
    r = secao(ws, r, "VERIFICACOES (NBR 6118, Tabela 13.3)")
    bal = f'{cel_vinc}="Balanco"'
    R["lim1"] = r; r = linha(ws, r, "Limite visual: L/250 (balanco L/125)", f"=IF({bal},{celvao_cm}/125,{celvao_cm}/250)", "cm")
    R["v1"] = r; r = linha(ws, r, "1) Aceitabilidade visual (flecha total)", f'=IF(B{R["atot"]}<=B{R["lim1"]},"OK","NAO PASSA")')
    R["avib"] = r; r = linha(ws, r, "Flecha da carga acidental", f"=IF({cel_plin}=0,0,{cel_aimed}*{cel_qlin}/{cel_plin})", "cm")
    R["lim2"] = r; r = linha(ws, r, "Limite vibracao: L/350 (balanco L/175)", f"=IF({bal},{celvao_cm}/175,{celvao_cm}/350)", "cm")
    R["v2"] = r; r = linha(ws, r, "2) Vibracoes sentidas (carga acidental)", f'=IF(B{R["avib"]}<=B{R["lim2"]},"OK","NAO PASSA")')
    R["apar"] = r; r = linha(ws, r, "Flecha apos construcao das paredes", f"=B{R['adif']}+IF({cel_plin}=0,0,{cel_aimed}*{cel_psi2}*{cel_qlin}/{cel_plin})", "cm", "diferida + imediata da acidental q.p.")
    R["lim3"] = r; r = linha(ws, r, "Limite paredes: min(L/500; 10 mm)", f"=MIN(IF({bal},{celvao_cm}/250,{celvao_cm}/500),1)", "cm")
    R["v3"] = r; r = linha(ws, r, "3) Paredes sobre o elemento", f'=IF({cel_parede}="Nao","N/A",IF(B{R["apar"]}<=B{R["lim3"]},"OK","NAO PASSA"))')
    r += 1
    R["status"] = r
    ws.cell(row=r, column=1, value="STATUS GERAL DO ELEMENTO").font = Font(bold=True, size=12)
    c = ws.cell(row=r, column=2, value=f'=IF(COUNTIF(B{R["v1"]}:B{R["v3"]},"NAO PASSA")>0,"NAO PASSA","OK")')
    c.font = Font(bold=True, size=12); c.border = borda
    cond_status(ws, f"B{R['v1']}:B{R['v3']}"); cond_status(ws, f"B{R['status']}")
    return r + 2, R


def secao_materiais(ws, r, R, cel_fck, cel_agreg):
    r = secao(ws, r, "MATERIAIS (NBR 6118, 8.2.5 e 8.2.8)")
    R["fctm"] = r; r = linha(ws, r, "fct,m = 0,3*fck^(2/3)", f"=0.3*{cel_fck}^(2/3)", "MPa")
    R["aE"] = r; r = linha(ws, r, "alfa_E (agregado)", F_AE.format(a=cel_agreg), "-")
    R["Eci"] = r; r = linha(ws, r, "Eci = alfa_E*5600*raiz(fck)", f"=B{R['aE']}*5600*SQRT({cel_fck})", "MPa", "valido p/ fck 20 a 50 MPa", fmt="0")
    R["ai"] = r; r = linha(ws, r, "alfa_i = 0,8 + 0,2*fck/80 <= 1", f"=MIN(0.8+0.2*{cel_fck}/80,1)", "-")
    R["Ecs"] = r; r = linha(ws, r, "Ecs = alfa_i*Eci", f"=B{R['ai']}*B{R['Eci']}/10", "kN/cm2", "nas flechas e convertido p/ tf/cm2")
    R["ae"] = r; r = linha(ws, r, "alfa_e = Es/Ecs (Es = 210 GPa)", f"=2100/B{R['Ecs']}*10", "-")
    return r + 1, R


def aba_retangular(ws, nome_elem, defaults, unid_carga, faixa_nota, b_fixo=None):
    """VIGA (b livre) ou LAJE MACICA (b=100)."""
    titulo(ws, nome_elem, "Flecha imediata (Branson) + diferida + verificacoes | NBR 6118:2023 itens 17.3.2.1.1, 17.3.2.1.2 e Tabela 13.3 | unidades: tf, cm")
    R = {}; r = 4
    r = secao(ws, r, "ENTRADAS (celulas amarelas)")
    R["L"] = r; r = linha(ws, r, "Vao L", defaults["L"], "m", "", "in", "0.00")
    R["vinc"] = r; r = linha(ws, r, "Vinculacao", defaults["vinc"], "", "Biapoiada / Apoiada-Engastada / Biengastada / Balanco", "in", "@")
    R["fck"] = r; r = linha(ws, r, "fck", defaults["fck"], "MPa", "", "in", "0")
    R["agreg"] = r; r = linha(ws, r, "Agregado graudo", defaults["agreg"], "", "afeta o modulo E", "in", "@")
    if b_fixo:
        R["b"] = r; r = linha(ws, r, "Largura da faixa b", b_fixo, "cm", "faixa de 1 m (fixa)", "calc", "0")
    else:
        R["b"] = r; r = linha(ws, r, "Largura b", defaults["b"], "cm", "", "in", "0.0")
    R["h"] = r; r = linha(ws, r, "Altura h", defaults["h"], "cm", "", "in", "0.0")
    R["d"] = r; r = linha(ws, r, "Altura util d", defaults["d"], "cm", "sugestao: h - 3 a 5 cm", "in", "0.0")
    R["dl"] = r; r = linha(ws, r, "d' (centro da armadura comprimida)", defaults["dl"], "cm", "", "in", "0.0")
    R["g"] = r; r = linha(ws, r, "Carga permanente g (INCLUIR peso proprio)", defaults["g"], unid_carga, faixa_nota, "in", "0.00")
    R["q"] = r; r = linha(ws, r, "Carga acidental q", defaults["q"], unid_carga, "", "in", "0.00")
    R["psi2"] = r; r = linha(ws, r, "psi_2 (quase-permanente)", defaults["psi2"], "-", "0,3 residencial | 0,4 comercial | 0,6 deposito", "in", "0.0")
    R["As"] = r; r = linha(ws, r, "As tracao" + defaults.get("as_nota", ""), defaults["As"], "cm2", defaults.get("as_nota2", ""), "in", "0.00")
    R["Asl"] = r; r = linha(ws, r, "As' compressao (montagem no vao)", defaults["Asl"], "cm2", "0 se nao houver", "in", "0.00")
    R["t0"] = r; r = linha(ws, r, "Idade do carregamento t0", defaults["t0"], "meses", "1 mes e o usual", "in", "0.0")
    R["parede"] = r; r = linha(ws, r, "Parede apoiada no elemento?", defaults["parede"], "", "Sim / Nao", "in", "@")
    r += 1
    dv_lista(ws, f"B{R['vinc']}", VINC); dv_lista(ws, f"B{R['agreg']}", AGREG); dv_lista(ws, f"B{R['parede']}", "Sim,Nao")

    r, R = secao_materiais(ws, r, R, f"B{R['fck']}", f"B{R['agreg']}")

    r = secao(ws, r, "SECAO BRUTA E MOMENTO DE FISSURACAO")
    R["Lcm"] = r; r = linha(ws, r, "Vao em cm", f"=B{R['L']}*100", "cm", fmt="0")
    R["Ic"] = r; r = linha(ws, r, "Ic = b*h^3/12", f"=B{R['b']}*B{R['h']}^3/12", "cm4", fmt="0")
    R["Mr"] = r; r = linha(ws, r, "Mr = 1,5*fct,m*Ic/yt  (yt = h/2)", f"=1.5*(B{R['fctm']}/100)*B{R['Ic']}/(B{R['h']}/2)", "tf.cm", "alfa=1,5 secao retangular", fmt="0.0")
    r += 1

    r = secao(ws, r, "CARGAS E ESFORCOS (combinacao quase-permanente)")
    if b_fixo:
        R["plin"] = r; r = linha(ws, r, "p = (g + psi2*q) x 1,00 m de faixa", f"=(B{R['g']}+B{R['psi2']}*B{R['q']})*1", "tf/m")
        R["qlin"] = r; r = linha(ws, r, "parcela acidental integral (p/ vibracao)", f"=B{R['q']}*1", "tf/m")
    else:
        R["plin"] = r; r = linha(ws, r, "p = g + psi2*q", f"=B{R['g']}+B{R['psi2']}*B{R['q']}", "tf/m")
        R["qlin"] = r; r = linha(ws, r, "parcela acidental integral (p/ vibracao)", f"=B{R['q']}", "tf/m")
    R["cm"] = r; r = linha(ws, r, "Coef. de momento", F_CM.format(v=f"B{R['vinc']}"), "-", "Ma = coef*p*L2", fmt="0.0000")
    R["Ma"] = r; r = linha(ws, r, "Ma (momento de servico)", f"=B{R['cm']}*(B{R['plin']}/100)*B{R['Lcm']}^2", "tf.cm", fmt="0.0")
    R["fiss"] = r; r = linha(ws, r, "Secao fissura? (Ma > Mr)", f'=IF(B{R["Ma"]}>B{R["Mr"]},"Sim - Estadio II","Nao - Estadio I")', "", "", "calc", "@")
    r += 1

    r = secao(ws, r, "ESTADIO II E RIGIDEZ EQUIVALENTE (Branson)")
    R["Aq"] = r; r = linha(ws, r, "aux A = b/2", f"=B{R['b']}/2", "")
    R["Bq"] = r; r = linha(ws, r, "aux B = (ae-1)As' + ae*As", f"=(B{R['ae']}-1)*B{R['Asl']}+B{R['ae']}*B{R['As']}", "")
    R["Cq"] = r; r = linha(ws, r, "aux C = (ae-1)As'd' + ae*As*d", f"=(B{R['ae']}-1)*B{R['Asl']}*B{R['dl']}+B{R['ae']}*B{R['As']}*B{R['d']}", "")
    R["xII"] = r; r = linha(ws, r, "Linha neutra x_II", f"=(-B{R['Bq']}+SQRT(B{R['Bq']}^2+4*B{R['Aq']}*B{R['Cq']}))/(2*B{R['Aq']})", "cm")
    R["III"] = r; r = linha(ws, r, "I_II (inercia fissurada)", f"=B{R['b']}*B{R['xII']}^3/3+(B{R['ae']}-1)*B{R['Asl']}*(B{R['xII']}-B{R['dl']})^2+B{R['ae']}*B{R['As']}*(B{R['d']}-B{R['xII']})^2", "cm4", fmt="0")
    R["Ieq"] = r; r = linha(ws, r, "I_eq = (Mr/Ma)3*Ic + [1-(Mr/Ma)3]*I_II <= Ic", f"=IF(B{R['Ma']}<=B{R['Mr']},B{R['Ic']},MIN(B{R['Ic']},(B{R['Mr']}/B{R['Ma']})^3*B{R['Ic']}+(1-(B{R['Mr']}/B{R['Ma']})^3)*B{R['III']}))", "cm4", fmt="0")
    r += 1

    r = secao(ws, r, "FLECHA IMEDIATA")
    R["k"] = r; r = linha(ws, r, "Coef. de flecha k", F_K.format(v=f"B{R['vinc']}"), "-", "5/384 biap. | 1/185 ap-eng | 1/384 bieng | 1/8 balanco", fmt="0.00000")
    R["aimed"] = r; r = linha(ws, r, "FLECHA IMEDIATA a_i = k*p*L^4/(Ecs*Ieq)", f"=B{R['k']}*(B{R['plin']}/100)*B{R['Lcm']}^4/((B{R['Ecs']}/10)*B{R['Ieq']})", "cm", "Ecs em tf/cm2")
    R["rho"] = r; r = linha(ws, r, "rho' = As'/(b*d)", f"=B{R['Asl']}/(B{R['b']}*B{R['d']})", "-", fmt="0.0000")
    r += 1

    r, R = bloco_flechas_verifs(ws, r, R, f"B{R['Lcm']}", f"B{R['aimed']}", f"B{R['qlin']}",
                                f"B{R['plin']}", f"B{R['psi2']}", f"B{R['t0']}", f"B{R['rho']}",
                                f"B{R['parede']}", f"B{R['vinc']}")
    return R


def aba_trelicada(ws):
    titulo(ws, "LAJE TRELICADA (vigota + capa, secao T por nervura)",
           "Flecha imediata (Branson, secao T) + diferida + verificacoes | NBR 6118:2023 | unidades: tf, cm | calculo por nervura")
    R = {}; r = 4
    r = secao(ws, r, "ENTRADAS (celulas amarelas)")
    R["L"] = r; r = linha(ws, r, "Vao L", 4.0, "m", "", "in", "0.00")
    R["vinc"] = r; r = linha(ws, r, "Vinculacao", "Biapoiada", "", "trelicada usual: Biapoiada", "in", "@")
    R["fck"] = r; r = linha(ws, r, "fck (capa/concreto local)", 25, "MPa", "", "in", "0")
    R["agreg"] = r; r = linha(ws, r, "Agregado graudo", "Granito/Gnaisse", "", "", "in", "@")
    R["bi"] = r; r = linha(ws, r, "Intereixo bi (largura da mesa)", 49.0, "cm", "vigota + tavela (ex.: 49)", "in", "0.0")
    R["bw"] = r; r = linha(ws, r, "Largura da nervura bw", 12.0, "cm", "base de concreto da vigota", "in", "0.0")
    R["h"] = r; r = linha(ws, r, "Altura total h (enchimento + capa)", 20.0, "cm", "ex.: 16+4 = 20", "in", "0.0")
    R["hf"] = r; r = linha(ws, r, "Espessura da capa hf", 4.0, "cm", "", "in", "0.0")
    R["d"] = r; r = linha(ws, r, "Altura util d", 17.0, "cm", "sugestao: h - 3", "in", "0.0")
    R["dl"] = r; r = linha(ws, r, "d' (banzo superior da trelica)", 1.5, "cm", "", "in", "0.0")
    R["g"] = r; r = linha(ws, r, "Carga permanente g (INCLUIR peso proprio)", 0.33, "tf/m2", "", "in", "0.00")
    R["q"] = r; r = linha(ws, r, "Carga acidental q", 0.15, "tf/m2", "", "in", "0.00")
    R["psi2"] = r; r = linha(ws, r, "psi_2 (quase-permanente)", 0.3, "-", "0,3 resid. | 0,4 comercial | 0,6 deposito", "in", "0.0")
    R["As"] = r; r = linha(ws, r, "As tracao POR NERVURA", 1.57, "cm2", "banzo inferior + adicional (ex.: 2 fi 10)", "in", "0.00")
    R["Asl"] = r; r = linha(ws, r, "As' POR NERVURA (banzo superior)", 0.28, "cm2", "1 fi 6 = 0,28", "in", "0.00")
    R["t0"] = r; r = linha(ws, r, "Idade do carregamento t0", 1.0, "meses", "", "in", "0.0")
    R["parede"] = r; r = linha(ws, r, "Parede apoiada na laje?", "Nao", "", "Sim / Nao", "in", "@")
    r += 1
    dv_lista(ws, f"B{R['vinc']}", VINC); dv_lista(ws, f"B{R['agreg']}", AGREG); dv_lista(ws, f"B{R['parede']}", "Sim,Nao")

    r, R = secao_materiais(ws, r, R, f"B{R['fck']}", f"B{R['agreg']}")

    r = secao(ws, r, "SECAO T BRUTA E MOMENTO DE FISSURACAO")
    R["Lcm"] = r; r = linha(ws, r, "Vao em cm", f"=B{R['L']}*100", "cm", fmt="0")
    R["A1"] = r; r = linha(ws, r, "Area da mesa A1 = bi*hf", f"=B{R['bi']}*B{R['hf']}", "cm2", fmt="0.0")
    R["A2"] = r; r = linha(ws, r, "Area da nervura A2 = bw*(h-hf)", f"=B{R['bw']}*(B{R['h']}-B{R['hf']})", "cm2", fmt="0.0")
    R["ycg"] = r; r = linha(ws, r, "Centroide (do topo) ycg", f"=(B{R['A1']}*B{R['hf']}/2+B{R['A2']}*(B{R['hf']}+(B{R['h']}-B{R['hf']})/2))/(B{R['A1']}+B{R['A2']})", "cm")
    R["Ic"] = r; r = linha(ws, r, "Ic (secao T bruta)", f"=B{R['bi']}*B{R['hf']}^3/12+B{R['A1']}*(B{R['ycg']}-B{R['hf']}/2)^2+B{R['bw']}*(B{R['h']}-B{R['hf']})^3/12+B{R['A2']}*(B{R['hf']}+(B{R['h']}-B{R['hf']})/2-B{R['ycg']})^2", "cm4", fmt="0")
    R["yt"] = r; r = linha(ws, r, "yt = h - ycg (fibra tracionada)", f"=B{R['h']}-B{R['ycg']}", "cm")
    R["Mr"] = r; r = linha(ws, r, "Mr = 1,2*fct,m*Ic/yt", f"=1.2*(B{R['fctm']}/100)*B{R['Ic']}/B{R['yt']}", "tf.cm", "alfa=1,2 secao T", fmt="0.0")
    r += 1

    r = secao(ws, r, "CARGAS E ESFORCOS POR NERVURA (combinacao quase-permanente)")
    R["plin"] = r; r = linha(ws, r, "p = (g + psi2*q)*bi/100", f"=(B{R['g']}+B{R['psi2']}*B{R['q']})*B{R['bi']}/100", "tf/m")
    R["qlin"] = r; r = linha(ws, r, "parcela acidental integral (p/ vibracao)", f"=B{R['q']}*B{R['bi']}/100", "tf/m")
    R["cm"] = r; r = linha(ws, r, "Coef. de momento", F_CM.format(v=f"B{R['vinc']}"), "-", fmt="0.0000")
    R["Ma"] = r; r = linha(ws, r, "Ma (momento de servico)", f"=B{R['cm']}*(B{R['plin']}/100)*B{R['Lcm']}^2", "tf.cm", fmt="0.0")
    R["fiss"] = r; r = linha(ws, r, "Secao fissura? (Ma > Mr)", f'=IF(B{R["Ma"]}>B{R["Mr"]},"Sim - Estadio II","Nao - Estadio I")', "", "", "calc", "@")
    r += 1

    r = secao(ws, r, "ESTADIO II E RIGIDEZ EQUIVALENTE (Branson, secao T)")
    R["B1"] = r; r = linha(ws, r, "aux B (retang. c/ b=bi)", f"=(B{R['ae']}-1)*B{R['Asl']}+B{R['ae']}*B{R['As']}", "")
    R["C1"] = r; r = linha(ws, r, "aux C (retang. c/ b=bi)", f"=(B{R['ae']}-1)*B{R['Asl']}*B{R['dl']}+B{R['ae']}*B{R['As']}*B{R['d']}", "")
    R["x1"] = r; r = linha(ws, r, "x_II supondo LN na capa", f"=(-B{R['B1']}+SQRT(B{R['B1']}^2+2*B{R['bi']}*B{R['C1']}))/B{R['bi']}", "cm")
    R["B2"] = r; r = linha(ws, r, "aux B (T completa)", f"=(B{R['bi']}-B{R['bw']})*B{R['hf']}+B{R['B1']}", "")
    R["C2"] = r; r = linha(ws, r, "aux C (T completa)", f"=(B{R['bi']}-B{R['bw']})*B{R['hf']}^2/2+B{R['C1']}", "")
    R["x2"] = r; r = linha(ws, r, "x_II supondo LN na nervura", f"=(-B{R['B2']}+SQRT(B{R['B2']}^2+2*B{R['bw']}*B{R['C2']}))/B{R['bw']}", "cm")
    R["xII"] = r; r = linha(ws, r, "x_II adotado", f"=IF(B{R['x1']}<=B{R['hf']},B{R['x1']},B{R['x2']})", "cm", "se <= hf a LN esta na capa (caso usual)")
    R["III"] = r; r = linha(ws, r, "I_II (inercia fissurada)",
        f"=IF(B{R['x1']}<=B{R['hf']},"
        f"B{R['bi']}*B{R['xII']}^3/3,"
        f"B{R['bw']}*B{R['xII']}^3/3+(B{R['bi']}-B{R['bw']})*B{R['hf']}^3/12+(B{R['bi']}-B{R['bw']})*B{R['hf']}*(B{R['xII']}-B{R['hf']}/2)^2)"
        f"+(B{R['ae']}-1)*B{R['Asl']}*(B{R['xII']}-B{R['dl']})^2+B{R['ae']}*B{R['As']}*(B{R['d']}-B{R['xII']})^2", "cm4", fmt="0")
    R["Ieq"] = r; r = linha(ws, r, "I_eq (Branson) <= Ic", f"=IF(B{R['Ma']}<=B{R['Mr']},B{R['Ic']},MIN(B{R['Ic']},(B{R['Mr']}/B{R['Ma']})^3*B{R['Ic']}+(1-(B{R['Mr']}/B{R['Ma']})^3)*B{R['III']}))", "cm4", fmt="0")
    r += 1

    r = secao(ws, r, "FLECHA IMEDIATA")
    R["k"] = r; r = linha(ws, r, "Coef. de flecha k", F_K.format(v=f"B{R['vinc']}"), "-", fmt="0.00000")
    R["aimed"] = r; r = linha(ws, r, "FLECHA IMEDIATA a_i = k*p*L^4/(Ecs*Ieq)", f"=B{R['k']}*(B{R['plin']}/100)*B{R['Lcm']}^4/((B{R['Ecs']}/10)*B{R['Ieq']})", "cm", "Ecs em tf/cm2")
    R["rho"] = r; r = linha(ws, r, "rho' = As'/(bi*d)", f"=B{R['Asl']}/(B{R['bi']}*B{R['d']})", "-", fmt="0.0000")
    r += 1

    r, R = bloco_flechas_verifs(ws, r, R, f"B{R['Lcm']}", f"B{R['aimed']}", f"B{R['qlin']}",
                                f"B{R['plin']}", f"B{R['psi2']}", f"B{R['t0']}", f"B{R['rho']}",
                                f"B{R['parede']}", f"B{R['vinc']}")
    return R


def aba_rapido(ws):
    titulo(ws, "MODO RAPIDO - flecha imediata vinda do software (Eberick/TQS)",
           "Cole a flecha imediata (secante) e o vao; a planilha calcula a diferida, a total e as verificacoes | NBR 6118, 17.3.2.1.2 e Tab. 13.3")
    R = {}; r = 4
    r = secao(ws, r, "ENTRADAS (celulas amarelas)")
    R["L"] = r; r = linha(ws, r, "Vao L", 5.71, "m", "", "in", "0.00")
    R["ai"] = r; r = linha(ws, r, "Flecha imediata do software", 0.381, "cm", "flecha 'secante' do Eberick", "in", "0.000")
    R["aq"] = r; r = linha(ws, r, "Flecha imediata SO da acidental (opcional)", 0.0, "cm", "0 = verificacao de vibracao nao avaliada", "in", "0.000")
    R["t0"] = r; r = linha(ws, r, "Idade do carregamento t0", 1.0, "meses", "1 mes e o usual", "in", "0.0")
    R["rho"] = r; r = linha(ws, r, "rho' de compressao no vao", 0.0, "%", "0 p/ laje macica; viga: As'/(b*d)*100", "in", "0.00")
    R["bal"] = r; r = linha(ws, r, "Elemento em balanco?", "Nao", "", "Sim / Nao", "in", "@")
    R["parede"] = r; r = linha(ws, r, "Parede apoiada no elemento?", "Sim", "", "Sim / Nao", "in", "@")
    r += 1
    dv_lista(ws, f"B{R['bal']}", "Sim,Nao"); dv_lista(ws, f"B{R['parede']}", "Sim,Nao")

    r = secao(ws, r, "RESULTADOS")
    R["Lcm"] = r; r = linha(ws, r, "Vao em cm", f"=B{R['L']}*100", "cm", fmt="0")
    R["xi0"] = r; r = linha(ws, r, "xi(t0)", f"=IF(B{R['t0']}>=70,2,0.68*(0.996^B{R['t0']})*(B{R['t0']}^0.32))", "-")
    R["af"] = r; r = linha(ws, r, "alfa_f = (2 - xi0)/(1 + 50*rho')", f"=(2-B{R['xi0']})/(1+50*B{R['rho']}/100)", "-")
    R["adif"] = r; r = linha(ws, r, "FLECHA DIFERIDA adicional", f"=B{R['ai']}*B{R['af']}", "cm")
    R["atot"] = r; r = linha(ws, r, "FLECHA TOTAL", f"=B{R['ai']}*(1+B{R['af']})", "cm")
    R["cx"] = r; r = linha(ws, r, "Relacao vao/flecha (C/x)", f"=IF(B{R['atot']}=0,0,B{R['Lcm']}/B{R['atot']})", "-", fmt="0")
    r += 1

    r = secao(ws, r, "VERIFICACOES (NBR 6118, Tabela 13.3)")
    bal = f'B{R["bal"]}="Sim"'
    R["lim1"] = r; r = linha(ws, r, "Limite visual L/250 (balanco L/125)", f"=IF({bal},B{R['Lcm']}/125,B{R['Lcm']}/250)", "cm")
    R["v1"] = r; r = linha(ws, r, "1) Aceitabilidade visual (flecha total)", f'=IF(B{R["atot"]}<=B{R["lim1"]},"OK","NAO PASSA")')
    R["lim2"] = r; r = linha(ws, r, "Limite vibracao L/350 (balanco L/175)", f"=IF({bal},B{R['Lcm']}/175,B{R['Lcm']}/350)", "cm")
    R["v2"] = r; r = linha(ws, r, "2) Vibracoes (se flecha da acidental informada)", f'=IF(B{R["aq"]}=0,"N/A",IF(B{R["aq"]}<=B{R["lim2"]},"OK","NAO PASSA"))')
    R["lim3"] = r; r = linha(ws, r, "Limite paredes min(L/500; 10 mm)", f"=MIN(IF({bal},B{R['Lcm']}/250,B{R['Lcm']}/500),1)", "cm")
    R["v3"] = r; r = linha(ws, r, "3) Paredes (flecha diferida apos alvenaria)", f'=IF(B{R["parede"]}="Nao","N/A",IF(B{R["adif"]}+B{R["aq"]}*0.3<=B{R["lim3"]},"OK","NAO PASSA"))')
    r += 1
    R["status"] = r
    ws.cell(row=r, column=1, value="STATUS GERAL").font = Font(bold=True, size=12)
    c = ws.cell(row=r, column=2, value=f'=IF(COUNTIF(B{R["v1"]}:B{R["v3"]},"NAO PASSA")>0,"NAO PASSA","OK")')
    c.font = Font(bold=True, size=12); c.border = borda
    cond_status(ws, f"B{R['v1']}:B{R['v3']}"); cond_status(ws, f"B{R['status']}")
    return R


def aba_dashboard(ws, refs):
    titulo(ws, "DASHBOARD - FLECHAS NBR 6118 | D'LIMA ENGENHARIA",
           "Resumo dos 4 modulos. Preencha as celulas AMARELAS de cada aba; tudo o mais calcula sozinho. NBR 6118:2023 - itens 17.3.2.1.1, 17.3.2.1.2 e Tabela 13.3")
    r = 4
    r = secao(ws, r, "RESUMO DOS ELEMENTOS")
    cab = ["Modulo (aba)", "Vao (m)", "Flecha total (cm)", "C/x", "Status"]
    for j, t in enumerate(cab, start=1):
        c = ws.cell(row=r, column=j, value=t)
        c.font = Font(bold=True); c.fill = PatternFill("solid", start_color=CINZA); c.border = borda
    r += 1
    for nome, aba, R in refs:
        ws.cell(row=r, column=1, value=nome).border = borda
        for j, key, fmt in ((2, "L", "0.00"), (3, "atot", "0.000"), (4, "cx", "0"), (5, "status", "@")):
            c = ws.cell(row=r, column=j, value=f"='{aba}'!B{R[key]}")
            c.font = F_LINK if j < 5 else Font(bold=True)
            c.number_format = fmt; c.border = borda
        r += 1
    cond_status(ws, f"E6:E{r-1}")
    r += 1
    r = secao(ws, r, "COMO USAR")
    notas = [
        "1) MODO RAPIDO: ja tem a flecha imediata do Eberick/TQS? Cole la o vao e a flecha; sai diferida, total e verificacoes.",
        "2) VIGA / LAJE MACICA / LAJE TRELICADA: calculo completo pela norma (Branson + estadio II + diferida) a partir de vao, secao, cargas e armadura.",
        "3) Celulas AMARELAS = voce preenche. Celulas CINZAS = calculadas (nao digite nelas).",
        "4) Verificacoes da Tabela 13.3: visual L/250, vibracao L/350 (so carga acidental), paredes min(L/500; 10 mm) sobre a flecha apos a alvenaria.",
        "5) Em 'g' inclua o peso proprio. CARGAS EM tf (1 tf = 10 kN). psi2: 0,3 residencial, 0,4 comercial, 0,6 deposito (NBR 6118, Tab. 11.2).",
        "6) Balanco: os limites ja trocam sozinhos (L/125, L/175, L/250) ao escolher 'Balanco' na vinculacao.",
        "7) Na trelicada, As' e o banzo superior da trelica (1 fi 6 = 0,28 cm2) e As inclui o banzo inferior + armadura adicional, por nervura.",
        "8) Conferencia do metodo: com vao 5,71 m e flecha 0,381 cm no MODO RAPIDO, a total da 0,885 cm (C/645) - mesmo caso validado no Eberick.",
    ]
    for n in notas:
        ws.merge_cells(f"A{r}:E{r}")
        ws.cell(row=r, column=1, value=n).font = Font(size=9)
        r += 1


def aba_norma(ws):
    titulo(ws, "BASE NORMATIVA E FORMULAS", "Referencias exatas da NBR 6118:2023 usadas na planilha")
    linhas = [
        ("", ""),
        ("ITEM 17.3.2.1.1 - Flecha imediata (rigidez equivalente de Branson):", ""),
        ("   (EI)eq = Ecs*{(Mr/Ma)^3*Ic + [1-(Mr/Ma)^3]*I_II} <= Ecs*Ic", ""),
        ("   Mr = alfa*fct,m*Ic/yt  com alfa = 1,5 (retangular) e 1,2 (secao T)", ""),
        ("ITEM 17.3.2.1.2 - Flecha diferida no tempo:", ""),
        ("   a_total = a_imediata*(1 + alfa_f)     alfa_f = (xi(t) - xi(t0)) / (1 + 50*rho')", ""),
        ("   xi(t) = 0,68*(0,996^t)*t^0,32 para t <= 70 meses;  xi(infinito) = 2,0  (t em meses)", ""),
        ("   Valores de xi (Tabela 17.1): t0=0,5 -> 0,54 | 1 -> 0,68 | 2 -> 0,84 | 4 -> 1,04 | 10 -> 1,36 | >=70 -> 2,0", ""),
        ("TABELA 13.3 - Limites de deslocamento:", ""),
        ("   Aceitabilidade sensorial visual: L/250 (balanco L/125) sobre a flecha total", ""),
        ("   Vibracoes sentidas no piso: L/350 (balanco L/175) sobre a flecha da carga acidental", ""),
        ("   Paredes: L/500 (balanco L/250) ou 10 mm, sobre a flecha ocorrida APOS a construcao da parede", ""),
        ("ITEM 8.2.5 - fct,m = 0,3*fck^(2/3) [MPa], fck 20 a 50 MPa", ""),
        ("ITEM 8.2.8 - Eci = alfa_E*5600*raiz(fck); Ecs = alfa_i*Eci; alfa_i = 0,8+0,2*fck/80 <= 1", ""),
        ("   alfa_E: basalto/diabasio 1,2 | granito/gnaisse 1,0 | calcario 0,9 | arenito 0,7", ""),
        ("ITEM 11.7.1 / TABELA 11.2 - psi2: 0,3 residencial | 0,4 comercial | 0,6 deposito", ""),
        ("", ""),
        ("Referencia: ABNT NBR 6118:2023 - Projeto de estruturas de concreto - Procedimento.", ""),
        ("Combinacao de servico usada: quase-permanente (g + psi2*q), adequada para ELS-DEF (item 11.8.3.1).", ""),
    ]
    r = 4
    for a, _ in linhas:
        ws.merge_cells(f"A{r}:E{r}")
        f = Font(bold=True, size=10) if (a and not a.startswith("   ")) else Font(size=9)
        ws.cell(row=r, column=1, value=a).font = f
        r += 1
    r += 1
    r = secao(ws, r, "NORMA DE CADA DADO DE ENTRADA (consulta rapida)")
    tabela = [
        ("Vao L (viga)", "NBR 6118 item 14.6.2.4 (vao efetivo lef = l0 + a1 + a2)"),
        ("Vinculacao / modelo de apoio", "NBR 6118 itens 14.6.4 e 14.6.7.1"),
        ("fck", "NBR 8953 (classes) | NBR 6118 Tab. 7.1 (CAA II: min C25)"),
        ("Agregado graudo (modulo E)", "NBR 6118 item 8.2.8 (alfa_E)"),
        ("Brita / dag", "NBR 7211 | NBR 6118 item 18.3.2.2 (espacamentos ah/av)"),
        ("Cobrimento nominal", "NBR 6118 Tab. 7.2 (por classe de agressividade)"),
        ("Bitola do estribo", "NBR 6118 item 18.3.3.2 (5 mm <= fi_t <= bw/10)"),
        ("Largura bw", "NBR 6118 item 13.2.2 (bw >= 12 cm; absoluto 10 cm)"),
        ("Altura h (pre-dim)", "pratica: h = L/10 a L/12"),
        ("Altura da laje macica", "NBR 6118 item 13.2.4.1 (min 8 cm laje de piso)"),
        ("psi2 (quase-permanente)", "NBR 6118 Tab. 11.2"),
        ("Idade de carregamento t0", "NBR 6118 item 17.3.2.1.2 e Tab. 17.1"),
        ("Sobrecarga de uso (q)", "NBR 6120:2019 Tab. 10 (lista pronta na aba LAJE-VIGA)"),
        ("Revestimento / contrapiso", "NBR 6120:2019 Tab. 1 a 3"),
        ("Peso de alvenaria", "NBR 6120:2019 Tab. 2 e 4"),
        ("Camadas / n de barras", "NBR 6118 item 18.3.2.2"),
        ("Bitolas comerciais", "NBR 7480"),
        ("Armadura minima As,min", "NBR 6118 item 17.3.5.2.1 e Tab. 17.3"),
        ("Ductilidade x/d <= 0,45", "NBR 6118 item 14.6.4.3"),
        ("Cisalhamento (modelo I)", "NBR 6118 item 17.4.2.2"),
        ("Limites de flecha", "NBR 6118 Tab. 13.3"),
        ("Combinacoes de servico", "NBR 6118 item 11.8.3.1"),
    ]
    for j, t in enumerate(("Dado / verificacao", "Onde consultar"), start=1):
        col = 1 if j == 1 else 3
        c = ws.cell(row=r, column=col, value=t)
        c.font = Font(bold=True, size=9); c.fill = PatternFill("solid", start_color=CINZA); c.border = borda
    r += 1
    for dado, ref in tabela:
        ws.cell(row=r, column=1, value=dado).border = borda
        ws.merge_cells(f"C{r}:E{r}")
        ws.cell(row=r, column=3, value=ref).border = borda
        ws.cell(row=r, column=3).font = Font(size=9)
        r += 1


# coeficientes de esforco por vinculacao (carga uniforme)
F_CMP = '=IF({v}="Biapoiada",1/8,IF({v}="Apoiada-Engastada",9/128,IF({v}="Biengastada",1/24,0)))'
F_CMN = '=IF({v}="Biapoiada",0,IF({v}="Apoiada-Engastada",1/8,IF({v}="Biengastada",1/12,1/2)))'
F_CV = '=IF({v}="Biapoiada",0.5,IF({v}="Apoiada-Engastada",0.625,IF({v}="Biengastada",0.5,1)))'
BARRAS = [(8, 0.503), (10, 0.785), (12.5, 1.227), (16, 2.011), (20, 3.142), (25, 4.909)]


FI_LISTA = "fi 8,fi 10,fi 12.5,fi 16,fi 20,fi 25"
F_AREA_FI = ('=IF({c}="fi 8",0.503,IF({c}="fi 10",0.785,IF({c}="fi 12.5",1.227,'
             'IF({c}="fi 16",2.011,IF({c}="fi 20",3.142,4.909)))))')
F_DIAM_FI = ('=IF({c}="fi 8",0.8,IF({c}="fi 10",1,IF({c}="fi 12.5",1.25,'
             'IF({c}="fi 16",1.6,IF({c}="fi 20",2,2.5)))))')


def aba_dimensionamento_viga(ws):
    """Viga retangular: ELU (flexao + cisalhamento) + ELS (flecha) na mesma aba,
    com carga automatica de laje macica, escolha de barras e desenho da secao. CA-50."""
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.marker import Marker

    titulo(ws, "DIMENSIONAMENTO DE VIGA - ELU + ELS | D'LIMA ENGENHARIA",
           "Cargas em tf, momentos em tf.m, geometria em cm | NBR 6118:2023 14.6.4.3, 17.2, 17.3.2, 17.4.2, Tab. 13.3 e 17.3 | NBR 6120:2019 | aco CA-50 | metodo validado c/ Musso e Carini")
    ws.column_dimensions["F"].width = 15
    R = {}; r = 4
    r = secao(ws, r, "ENTRADAS - VIGA (celulas amarelas)")
    R["L"] = r; r = linha(ws, r, "Vao L (balanco: comprimento do balanco)", 4.0, "m", "", "in", "0.00")
    R["vinc"] = r; r = linha(ws, r, "Vinculacao", "Biapoiada", "", "Biapoiada (ap-ap) | Apoiada-Engastada | Biengastada (eng-eng) | Balanco (eng-livre)", "in", "@")
    R["fck"] = r; r = linha(ws, r, "fck", 25, "MPa", "20 a 50 MPa", "in", "0")
    R["agreg"] = r; r = linha(ws, r, "Agregado graudo", "Granito/Gnaisse", "", "afeta o modulo E (flecha)", "in", "@")
    R["brita"] = r; r = linha(ws, r, "Brita", "Brita 1", "", "afeta espacamento minimo entre barras (dag)", "in", "@")
    R["dag"] = r; r = linha(ws, r, "Dimensao maxima do agregado dag", f'=IF(B{r-1}="Brita 0",0.95,IF(B{r-1}="Brita 1",1.9,2.5))', "cm", "brita 0 = 9,5 | brita 1 = 19 | brita 2 = 25 mm")
    R["cob"] = r; r = linha(ws, r, "Cobrimento nominal c", 2.5, "cm", "CAA I: 2,5 | CAA II: 3,0 | CAA III: 4,0 (NBR 6118 Tab. 7.2)", "in", "0.0")
    R["fie"] = r; r = linha(ws, r, "Bitola do estribo (2 ramos)", "fi 5", "", "escolha na lista; 5 mm <= fi_t <= bw/10", "in", "@")
    R["die"] = r; r = linha(ws, r, "Diametro do estribo", f'=IF(B{r-1}="fi 5",0.5,IF(B{r-1}="fi 6.3",0.63,0.8))', "cm")
    R["bw"] = r; r = linha(ws, r, "Largura bw", 14.0, "cm", "", "in", "0.0")
    R["h"] = r; r = linha(ws, r, "Altura h", 40.0, "cm", "d util e calculado sozinho das camadas escolhidas", "in", "0.0")
    R["psi2"] = r; r = linha(ws, r, "psi_2 (quase-permanente, p/ flecha)", 0.3, "-", "0,3 residencial | 0,4 comercial | 0,6 deposito", "in", "0.0")
    R["t0"] = r; r = linha(ws, r, "Idade do carregamento t0", 1.0, "meses", "1 mes e o usual", "in", "0.0")
    R["parede"] = r; r = linha(ws, r, "Parede apoiada na viga?", "Sim", "", "Sim / Nao", "in", "@")
    r += 1
    r = secao(ws, r, "ENTRADAS - LAJE MACICA APOIADA NA VIGA")
    R["hl"] = r; r = linha(ws, r, "Altura da laje h_laje", 10.0, "cm", "0 = sem laje", "in", "0.0")
    R["Ainf"] = r; r = linha(ws, r, "Area de influencia da laje", 9.0, "m2", "use a aba LAJE-VIGA p/ calcular pelo quinhao (45 graus)", "in", "0.00")
    R["grev"] = r; r = linha(ws, r, "Revestimento + contrapiso da laje", 0.10, "tf/m2", "usual: 0,10 (NBR 6120)", "in", "0.00")
    R["qlaje"] = r; r = linha(ws, r, "Carga acidental da laje", 0.15, "tf/m2", "veja a tabela de sobrecargas NBR 6120 na aba LAJE-VIGA", "in", "0.00")
    r += 1
    r = secao(ws, r, "ENTRADAS - OUTRAS CARGAS DIRETAS NA VIGA")
    R["galv"] = r; r = linha(ws, r, "Alvenaria + permanentes na viga", 0.36, "tf/m", "ex.: parede 14 cm x 2,8 m ~ 0,36", "in", "0.00")
    R["qout"] = r; r = linha(ws, r, "Acidentais diretas na viga", 0.0, "tf/m", "0 se nao houver", "in", "0.00")
    r += 1
    dv_lista(ws, f"B{R['vinc']}", VINC); dv_lista(ws, f"B{R['agreg']}", AGREG); dv_lista(ws, f"B{R['parede']}", "Sim,Nao")
    dv_lista(ws, f"B{R['brita']}", "Brita 0,Brita 1,Brita 2")
    dv_lista(ws, f"B{R['fie']}", "fi 5,fi 6.3,fi 8")

    ws.column_dimensions["E"].width = 36
    c = ws.cell(row=3, column=5, value="NORMA P/ CONSULTA RAPIDA")
    c.font = Font(bold=True, size=9, color="1A3C6E")
    NORMAS_IN = {
        "L": "NBR 6118 item 14.6.2.4 (vao efetivo lef)",
        "vinc": "NBR 6118 itens 14.6.4 e 14.6.7.1 (modelo de apoio)",
        "fck": "NBR 8953 (classes) | NBR 6118 Tab. 7.1 (CAA II: min C25)",
        "agreg": "NBR 6118 item 8.2.8 (alfa_E do modulo E)",
        "brita": "NBR 7211 | NBR 6118 item 18.3.2.2 (dag nos espacamentos)",
        "cob": "NBR 6118 Tab. 7.2 (cobrimento por CAA)",
        "fie": "NBR 6118 item 18.3.3.2 (5 mm <= fi_t <= bw/10)",
        "bw": "NBR 6118 item 13.2.2 (bw >= 12 cm; abs. 10 cm)",
        "h": "pratica de pre-dim: h = L/10 a L/12",
        "psi2": "NBR 6118 Tab. 11.2 (psi2 por uso)",
        "t0": "NBR 6118 item 17.3.2.1.2 e Tab. 17.1 (xi x tempo)",
        "parede": "NBR 6118 Tab. 13.3 (limite p/ alvenarias)",
        "hl": "NBR 6118 item 13.2.4.1 (h min: 8 cm laje de piso)",
        "Ainf": "charneiras a 45 graus - calcule na aba LAJE-VIGA",
        "grev": "NBR 6120:2019 Tab. 1 a 3 (revestimentos ~0,10 tf/m2)",
        "qlaje": "NBR 6120:2019 Tab. 10 (lista na aba LAJE-VIGA)",
        "galv": "NBR 6120:2019 Tab. 2 e 4 (peso de paredes)",
        "qout": "NBR 6120:2019 Tab. 10",
    }
    for k, ref in NORMAS_IN.items():
        ws.cell(row=R[k], column=5, value=ref).font = F_NOTA

    r = secao(ws, r, "CARGAS (automaticas, em tf)")
    R["ppv"] = r; r = linha(ws, r, "Peso proprio da viga = 2,5*bw*h", f"=2.5*B{R['bw']}*B{R['h']}/10000", "tf/m")
    R["ppl"] = r; r = linha(ws, r, "Peso proprio da laje = 2,5*h_laje", f"=2.5*B{R['hl']}/100", "tf/m2")
    R["gl"] = r; r = linha(ws, r, "g da laje na viga = (pp+rev)*Ainf/L", f"=(B{R['ppl']}+IF(B{R['hl']}=0,0,B{R['grev']}))*B{R['Ainf']}/B{R['L']}", "tf/m")
    R["ql"] = r; r = linha(ws, r, "q da laje na viga = q*Ainf/L", f"=IF(B{R['hl']}=0,0,B{R['qlaje']}*B{R['Ainf']}/B{R['L']})", "tf/m")
    R["g"] = r; r = linha(ws, r, "g TOTAL (pp viga + laje + alvenaria)", f"=B{R['ppv']}+B{R['gl']}+B{R['galv']}", "tf/m", "servico, sem majorar")
    R["q"] = r; r = linha(ws, r, "q TOTAL", f"=B{R['ql']}+B{R['qout']}", "tf/m", "servico, sem majorar")
    R["pd"] = r; r = linha(ws, r, "pd = 1,4*(g + q)  (ELU)", f"=1.4*(B{R['g']}+B{R['q']})", "tf/m")
    R["pserv"] = r; r = linha(ws, r, "p quase-permanente = g + psi2*q (ELS)", f"=B{R['g']}+B{R['psi2']}*B{R['q']}", "tf/m")
    r += 1

    r = secao(ws, r, "MATERIAIS DE CALCULO ELU (gama_c=1,4 | gama_s=1,15)")
    R["fcd"] = r; r = linha(ws, r, "fcd = fck/1,4", f"=B{R['fck']}/140", "tf/cm2")
    R["fctm2"] = r; r = linha(ws, r, "fct,m = 0,3*fck^(2/3)", f"=0.3*B{R['fck']}^(2/3)", "MPa")
    R["fctd"] = r; r = linha(ws, r, "fctd = 0,7*fct,m/1,4", f"=0.7*B{R['fctm2']}/140", "tf/cm2")
    R["fyd"] = r; r = linha(ws, r, "fyd = 50/1,15 (CA-50)", "=50/11.5", "tf/cm2")
    R["rmin"] = r; r = linha(ws, r, "rho_min (Tab. 17.3)", f"=IF(B{R['fck']}<=30,0.0015,IF(B{R['fck']}<=35,0.00164,IF(B{R['fck']}<=40,0.00179,IF(B{R['fck']}<=45,0.00194,0.00208))))", "-", "As_min = rho_min*bw*h", fmt="0.00000")
    R["Asmin"] = r; r = linha(ws, r, "As_min = rho_min*bw*h", f"=B{R['rmin']}*B{R['bw']}*B{R['h']}", "cm2")
    r += 1

    r = secao(ws, r, "ESFORCOS DE CALCULO (por vinculacao)")
    R["Lcm"] = r; r = linha(ws, r, "Vao em cm", f"=B{R['L']}*100", "cm", fmt="0")
    R["cmp"] = r; r = linha(ws, r, "Coef. momento POSITIVO (vao)", F_CMP.format(v=f"B{R['vinc']}"), "-", "1/8 biap | 9/128 ap-eng | 1/24 bieng | 0 balanco", fmt="0.0000")
    R["cmn"] = r; r = linha(ws, r, "Coef. momento NEGATIVO (apoio)", F_CMN.format(v=f"B{R['vinc']}"), "-", "0 biap | 1/8 ap-eng | 1/12 bieng | 1/2 balanco", fmt="0.0000")
    R["cv"] = r; r = linha(ws, r, "Coef. cortante", F_CV.format(v=f"B{R['vinc']}"), "-", "0,5 biap/bieng | 0,625 ap-eng (lado engaste) | 1 balanco", fmt="0.000")
    R["Mdp"] = r; r = linha(ws, r, "Md+ (vao)", f"=B{R['cmp']}*B{R['pd']}*B{R['L']}^2", "tf.m", fmt="0.00")
    R["Mdn"] = r; r = linha(ws, r, "Md- (apoio engastado)", f"=B{R['cmn']}*B{R['pd']}*B{R['L']}^2", "tf.m", fmt="0.00")
    R["Vd"] = r; r = linha(ws, r, "Vd (cortante maximo)", f"=B{R['cv']}*B{R['pd']}*B{R['L']}", "tf", fmt="0.00")
    r += 1

    def bloco_escolha(r, nome, kn, n_def, fi_def):
        r = secao(ws, r, f"{nome} - escolha o arranjo (dropdowns)")
        R[kn + "nc"] = r; r = linha(ws, r, "Quantas camadas?", 1, "un", "2a e 3a camadas tem SEMPRE 2 barras, nos cantos do estribo", "in", "0")
        R[kn] = r; r = linha(ws, r, "Barras na 1a camada", n_def, "un", "so a 1a camada pode ter mais de 2 (toda barra amarrada no estribo)", "in", "0")
        R[kn + "fi"] = r; r = linha(ws, r, "Bitola da 1a camada", fi_def, "", "escolha na lista", "in", "@")
        R[kn + "fi2"] = r; r = linha(ws, r, "Bitola da 2a camada (2 barras)", fi_def, "", "so e usada se houver 2a camada", "in", "@")
        R[kn + "fi3"] = r; r = linha(ws, r, "Bitola da 3a camada (2 barras)", fi_def, "", "so e usada se houver 3a camada", "in", "@")
        rd1 = r; r = linha(ws, r, "Diametro 1a camada", F_DIAM_FI.format(c=f"B{R[kn + 'fi']}"), "cm")
        rd2 = r; r = linha(ws, r, "Diametro 2a camada", F_DIAM_FI.format(c=f"B{R[kn + 'fi2']}"), "cm")
        rd3 = r; r = linha(ws, r, "Diametro 3a camada", F_DIAM_FI.format(c=f"B{R[kn + 'fi3']}"), "cm")
        a1 = F_AREA_FI.format(c=f"B{R[kn + 'fi']}")[1:]
        a2 = F_AREA_FI.format(c=f"B{R[kn + 'fi2']}")[1:]
        a3 = F_AREA_FI.format(c=f"B{R[kn + 'fi3']}")[1:]
        ra1 = r; r = linha(ws, r, "As da 1a camada = n1 x area", f"=B{R[kn]}*({a1})", "cm2")
        ra2 = r; r = linha(ws, r, "As da 2a camada (2 barras)", f"=IF(B{R[kn + 'nc']}>=2,2*({a2}),0)", "cm2")
        ra3 = r; r = linha(ws, r, "As da 3a camada (2 barras)", f"=IF(B{R[kn + 'nc']}>=3,2*({a3}),0)", "cm2")
        R[kn + "ef"] = r; r = linha(ws, r, "As EFETIVO TOTAL", f"=B{ra1}+B{ra2}+B{ra3}", "cm2")
        ws.cell(row=R[kn + "ef"], column=1).font = Font(bold=True)
        ws.cell(row=R[kn + "ef"], column=2).font = Font(bold=True)
        R[kn + "y1"] = r; r = linha(ws, r, "Centro da 1a camada ate a face", f"=B{R['cob']}+B{R['die']}+B{rd1}/2", "cm", "cobrimento + estribo + fi/2 (Musso: dc = ct + fi_t + fi/2)")
        R[kn + "y2"] = r; r = linha(ws, r, "Centro da 2a camada ate a face", f"=IF(B{R[kn + 'nc']}>=2,B{R[kn + 'y1']}+B{rd1}/2+MAX(2,B{rd1},B{rd2},0.5*B{R['dag']})+B{rd2}/2,B{R[kn + 'y1']})", "cm", "esp. livre av >= max(2 cm; fi; 0,5*dag) - NBR 6118 18.3.2.2")
        R[kn + "y3"] = r; r = linha(ws, r, "Centro da 3a camada ate a face", f"=IF(B{R[kn + 'nc']}>=3,B{R[kn + 'y2']}+B{rd2}/2+MAX(2,B{rd2},B{rd3},0.5*B{R['dag']})+B{rd3}/2,B{R[kn + 'y2']})", "cm")
        rycg = r; r = linha(ws, r, "Centroide da armadura ate a face (dc)", f"=(B{ra1}*B{R[kn + 'y1']}+B{ra2}*B{R[kn + 'y2']}+B{ra3}*B{R[kn + 'y3']})/B{R[kn + 'ef']}", "cm")
        R[kn + "d"] = r; r = linha(ws, r, "d UTIL CALCULADO = h - centroide", f"=B{R['h']}-B{rycg}", "cm")
        ws.cell(row=R[kn + "d"], column=1).font = Font(bold=True)
        ws.cell(row=R[kn + "d"], column=2).font = Font(bold=True)
        R[kn + "dc"] = r; r = linha(ws, r, "Centroide dentro de 10% de h?",
                                    f'=IF(B{rycg}<=0.1*B{R["h"]},"OK","NAO PASSA")', "", "criterio do d real (Musso): camadas demais afastam o centroide", "calc", "@")
        rlarg = r; r = linha(ws, r, "Largura necessaria (1a camada)", f"=2*(B{R['cob']}+B{R['die']})+B{R[kn]}*B{rd1}+(B{R[kn]}-1)*MAX(2,B{rd1},1.2*B{R['dag']})", "cm", "esp. livre ah >= max(2 cm; fi; 1,2*dag) - NBR 6118 18.3.2.2")
        R[kn + "cam"] = r; r = linha(ws, r, "Cabe na largura bw?",
                                     f'=IF(B{rlarg}<=B{R["bw"]},"OK","NAO PASSA")', "", "NAO PASSA: menos barras na 1a camada, bitola maior ou mais camadas", "calc", "@")
        R[kn + "chk"] = r; r = linha(ws, r, "Verificacao As efetivo >= As necessario", "", "", "compara com o As necessario da flexao abaixo", "calc", "@")
        dv_lista(ws, f"B{R[kn]}", "2,3,4,5,6,7,8")
        dv_lista(ws, f"B{R[kn + 'nc']}", "1,2,3")
        for suf in ("fi", "fi2", "fi3"):
            dv_lista(ws, f"B{R[kn + suf]}", FI_LISTA)
        ws.cell(row=R[kn + "nc"], column=5, value="NBR 6118 item 18.3.2.2 (camadas e av)").font = F_NOTA
        ws.cell(row=R[kn], column=5, value="NBR 6118 item 18.3.2.2 (ah; min 2 barras)").font = F_NOTA
        for suf in ("fi", "fi2", "fi3"):
            ws.cell(row=R[kn + suf], column=5, value="NBR 7480 (bitolas comerciais CA-50)").font = F_NOTA
        return r + 1

    r = bloco_escolha(r, "ARMADURA POSITIVA (vao, embaixo)", "nv", 2, "fi 16")
    r = bloco_escolha(r, "ARMADURA NEGATIVA (apoio, em cima)", "na", 2, "fi 10")

    def bloco_flexao(r, nome, cel_md, chk, dcel):
        r = secao(ws, r, f"FLEXAO ELU - {nome}")
        rx = r; r = linha(ws, r, "Linha neutra x = 1,25*d*(1-RAIZ(1-Md/(0,425*bw*d2*fcd)))",
                          f"=IF({cel_md}=0,0,1.25*{dcel}*(1-SQRT(MAX(0,1-{cel_md}*100/(0.425*B{R['bw']}*{dcel}^2*B{R['fcd']})))))", "cm")
        rxd = r; r = linha(ws, r, "x/d", f"=B{rx}/{dcel}", "-", "limite 0,45 p/ fck<=50 (14.6.4.3)")
        R[chk] = r; r = linha(ws, r, "Verificacao de ductilidade x/d <= 0,45",
                              f'=IF(B{rxd}<=0.45,"OK","NAO PASSA")', "", "NAO PASSA: aumente h/bw ou fck", "calc", "@")
        rc = r; r = linha(ws, r, "As calculado = Md/(fyd*(d-0,4x))",
                          f"=IF({cel_md}=0,0,{cel_md}*100/(B{R['fyd']}*({dcel}-0.4*B{rx})))", "cm2")
        radot = r; r = linha(ws, r, "As NECESSARIO = max(As calc; As_min)",
                             f"=IF({cel_md}=0,0,MAX(B{rc},B{R['Asmin']}))", "cm2",
                             "0 = sem momento: use armadura construtiva (2 fi 8)")
        ws.cell(row=radot, column=1).font = Font(bold=True)
        ws.cell(row=radot, column=2).font = Font(bold=True)
        return r + 1, radot

    r, R["Asp"] = bloco_flexao(r, "ARMADURA POSITIVA (vao)", f"B{R['Mdp']}", "chkp", f"B{R['nvd']}")
    r, R["Asn"] = bloco_flexao(r, "ARMADURA NEGATIVA (apoio)", f"B{R['Mdn']}", "chkn", f"B{R['nad']}")

    for kn, radot in (("nv", R["Asp"]), ("na", R["Asn"])):
        ws.cell(row=R[kn + "chk"], column=2,
                value=f'=IF(B{radot}=0,"OK",IF(B{R[kn + "ef"]}>=B{radot},"OK","NAO PASSA"))')

    r = secao(ws, r, "REFERENCIA: quantas barras de cada bitola atendem o As necessario")
    cab = ["Bitola (mm)", "Area/barra (cm2)", "n barras POSITIVA", "As efetivo POSITIVA", "n barras NEGATIVA", "As efetivo NEGATIVA"]
    for j, t in enumerate(cab, start=1):
        c = ws.cell(row=r, column=j, value=t)
        c.font = Font(bold=True, size=9); c.fill = PatternFill("solid", start_color=CINZA); c.border = borda
    r += 1
    for fi, area in BARRAS:
        ws.cell(row=r, column=1, value=fi).border = borda
        c = ws.cell(row=r, column=2, value=area); c.border = borda; c.number_format = "0.000"
        for jn, jef, radot in ((3, 4, R["Asp"]), (5, 6, R["Asn"])):
            cn = ws.cell(row=r, column=jn, value=f'=IF(B{radot}=0,"-",MAX(2,CEILING(B{radot}/B{r},1)))')
            cn.border = borda; cn.number_format = "0"
            ce = ws.cell(row=r, column=jef, value=f'=IF(B{radot}=0,"-",{chr(64+jn)}{r}*B{r})')
            ce.border = borda; ce.number_format = "0.00"
        r += 1
    r += 1

    r = secao(ws, r, "CISALHAMENTO (modelo de calculo I, item 17.4.2.2)")
    R["dtr"] = r; r = linha(ws, r, "d utilizado (lado tracionado)", f'=IF(B{R["vinc"]}="Balanco",B{R["nad"]},B{R["nvd"]})', "cm", "balanco traciona em cima: usa o d da armadura negativa")
    R["av2"] = r; r = linha(ws, r, "alfa_v2 = 1 - fck/250", f"=1-B{R['fck']}/250", "-")
    R["VRd2"] = r; r = linha(ws, r, "VRd2 = 0,27*alfa_v2*fcd*bw*d (biela)", f"=0.27*B{R['av2']}*B{R['fcd']}*B{R['bw']}*B{R['dtr']}", "tf", fmt="0.00")
    R["chkv"] = r; r = linha(ws, r, "Verificacao da biela Vd <= VRd2", f'=IF(B{R["Vd"]}<=B{R["VRd2"]},"OK","NAO PASSA")', "", "NAO PASSA: aumente a secao", "calc", "@")
    R["Vc"] = r; r = linha(ws, r, "Vc = 0,6*fctd*bw*d", f"=0.6*B{R['fctd']}*B{R['bw']}*B{R['dtr']}", "tf", fmt="0.00")
    R["aswc"] = r; r = linha(ws, r, "Asw/s calculado = (Vd-Vc)/(0,9*d*fyd)", f"=MAX(0,(B{R['Vd']}-B{R['Vc']})/(0.9*B{R['dtr']}*B{R['fyd']}))*100", "cm2/m")
    R["aswm"] = r; r = linha(ws, r, "Asw/s minimo = 0,2*(fct,m/500)*bw", f"=0.2*B{R['fctm2']}/500*B{R['bw']}*100", "cm2/m")
    R["asw"] = r; r = linha(ws, r, "Asw/s ADOTADO", f"=MAX(B{R['aswc']},B{R['aswm']})", "cm2/m")
    R["smax"] = r; r = linha(ws, r, "Espacamento maximo s_max", f"=IF(B{R['Vd']}<=0.67*B{R['VRd2']},MIN(0.6*B{R['dtr']},30),MIN(0.3*B{R['dtr']},20))", "cm", "0,6d<=30 se Vd<=0,67VRd2; senao 0,3d<=20", fmt="0.0")
    rare = r; r = linha(ws, r, "Area por ramo do estribo", '=IF(B{0}="fi 5",0.196,IF(B{0}="fi 6.3",0.312,0.503))'.format(R["fie"]), "cm2")
    R["sest"] = r; r = linha(ws, r, "ESTRIBO ADOTADO: espacamento s", f"=MIN(FLOOR(2*B{rare}/B{R['asw']}*100,1),FLOOR(B{R['smax']},1))", "cm", "ja limitado a s_max; bitola escolhida nas ENTRADAS", fmt="0")
    ws.cell(row=R["sest"], column=1).font = Font(bold=True); ws.cell(row=R["sest"], column=2).font = Font(bold=True)
    r += 1

    r = secao(ws, r, "ELS - FLECHA (mesma viga, com o As efetivo escolhido)")
    r, R = secao_materiais(ws, r, R, f"B{R['fck']}", f"B{R['agreg']}")
    R["Ic"] = r; r = linha(ws, r, "Ic = bw*h^3/12", f"=B{R['bw']}*B{R['h']}^3/12", "cm4", fmt="0")
    R["Mr"] = r; r = linha(ws, r, "Mr = 1,5*fct,m*Ic/(h/2)", f"=1.5*(B{R['fctm']}/100)*B{R['Ic']}/(B{R['h']}/2)", "tf.cm", "alfa=1,5 secao retangular", fmt="0.0")
    R["cme"] = r; r = linha(ws, r, "Coef. de momento (servico)", F_CM.format(v=f"B{R['vinc']}"), "-", fmt="0.0000")
    R["Ma"] = r; r = linha(ws, r, "Ma (momento quase-permanente)", f"=B{R['cme']}*(B{R['pserv']}/100)*B{R['Lcm']}^2", "tf.cm", fmt="0.0")
    R["fiss"] = r; r = linha(ws, r, "Secao fissura? (Ma > Mr)", f'=IF(B{R["Ma"]}>B{R["Mr"]},"Sim - Estadio II","Nao - Estadio I")', "", "", "calc", "@")
    R["Astr"] = r; r = linha(ws, r, "As de tracao usado (balanco: apoio)", f'=IF(B{R["vinc"]}="Balanco",B{R["naef"]},B{R["nvef"]})', "cm2", "As' desprezado (a favor da seguranca)")
    R["Bq"] = r; r = linha(ws, r, "aux B = ae*As", f"=B{R['ae']}*B{R['Astr']}", "")
    R["xII"] = r; r = linha(ws, r, "Linha neutra x_II (servico)", f"=(-B{R['Bq']}+SQRT(B{R['Bq']}^2+2*B{R['bw']}*B{R['Bq']}*B{R['dtr']}))/B{R['bw']}", "cm")
    R["III"] = r; r = linha(ws, r, "I_II (inercia fissurada)", f"=B{R['bw']}*B{R['xII']}^3/3+B{R['ae']}*B{R['Astr']}*(B{R['dtr']}-B{R['xII']})^2", "cm4", fmt="0")
    R["Ieq"] = r; r = linha(ws, r, "I_eq (Branson) <= Ic", f"=IF(B{R['Ma']}<=B{R['Mr']},B{R['Ic']},MIN(B{R['Ic']},(B{R['Mr']}/B{R['Ma']})^3*B{R['Ic']}+(1-(B{R['Mr']}/B{R['Ma']})^3)*B{R['III']}))", "cm4", fmt="0")
    R["k"] = r; r = linha(ws, r, "Coef. de flecha k", F_K.format(v=f"B{R['vinc']}"), "-", "5/384 biap | 1/185 ap-eng | 1/384 bieng | 1/8 balanco", fmt="0.00000")
    R["aimed"] = r; r = linha(ws, r, "FLECHA IMEDIATA a_i = k*p*L^4/(Ecs*Ieq)", f"=B{R['k']}*(B{R['pserv']}/100)*B{R['Lcm']}^4/((B{R['Ecs']}/10)*B{R['Ieq']})", "cm", "Ecs convertido p/ tf/cm2")
    r += 1

    r, R = bloco_flechas_verifs(ws, r, R, f"B{R['Lcm']}", f"B{R['aimed']}", f"B{R['q']}",
                                f"B{R['pserv']}", f"B{R['psi2']}", f"B{R['t0']}", "0",
                                f"B{R['parede']}", f"B{R['vinc']}")

    R["statusg"] = r
    ws.cell(row=r, column=1, value="STATUS GERAL: ELU + BARRAS + FLECHA").font = Font(bold=True, size=12)
    chks = [f'COUNTIF(B{R[k]},"NAO PASSA")' for k in
            ("chkp", "chkn", "chkv", "nvchk", "nvcam", "nvdc", "nachk", "nacam", "nadc", "status")]
    c = ws.cell(row=r, column=2, value=f'=IF({"+".join(chks)}>0,"NAO PASSA","OK")')
    c.font = Font(bold=True, size=12); c.border = borda
    for k in ("chkp", "chkn", "chkv", "nvchk", "nvcam", "nvdc", "nachk", "nacam", "nadc", "statusg"):
        cond_status(ws, f"B{R[k]}")
    r += 2
    notas = [
        "COMO USAR: preencha so as celulas AMARELAS (as demais sao protegidas; senha p/ desbloquear: dlima). Unidades: tf, tf/m, tf.m.",
        "Area de influencia: calcule na aba LAJE-VIGA (quinhao a 45 graus a partir da metragem da laje) e copie o valor aqui.",
        "Arranjo das armaduras: escolha camadas, barras da 1a camada e a bitola de CADA camada. 2a e 3a camadas tem sempre 2 barras nos cantos do estribo (toda barra amarrada).",
        "O d util e CALCULADO do arranjo (h - centroide das camadas), com espacamentos ah/av da NBR 6118 18.3.2.2 (inclui a brita) e criterio do centroide <= 10% h (Musso).",
        "A flecha (ELS) ja usa o As efetivo escolhido, com As' = 0 (a favor da seguranca) e combinacao quase-permanente.",
        "Momentos de apoio: em portico real o engaste e parcial; os coeficientes aqui sao os classicos de viga isolada.",
        "Aco CA-50. Para CA-60 em estribos, o espacamento adotado fica a favor da seguranca.",
    ]
    for n in notas:
        ws.merge_cells(f"A{r}:F{r}")
        ws.cell(row=r, column=1, value=n).font = Font(size=9)
        r += 1

    # ----- desenho da secao (grafico de dispersao alimentado por celulas auxiliares S/T) -----
    S, T = 19, 20
    ws.cell(row=4, column=S, value="aux desenho (nao apagar)").font = F_NOTA
    bw, h = f"B{R['bw']}", f"B{R['h']}"
    contorno = [(f"=-{bw}/2", f"=-{h}/2"), (f"={bw}/2", f"=-{h}/2"), (f"={bw}/2", f"={h}/2"),
                (f"=-{bw}/2", f"={h}/2"), (f"=-{bw}/2", f"=-{h}/2")]
    for i, (x, y) in enumerate(contorno):
        ws.cell(row=5 + i, column=S, value=x); ws.cell(row=5 + i, column=T, value=y)
    estribo = [(f"=-{bw}/2+2.5", f"=-{h}/2+2.5"), (f"={bw}/2-2.5", f"=-{h}/2+2.5"), (f"={bw}/2-2.5", f"={h}/2-2.5"),
               (f"=-{bw}/2+2.5", f"={h}/2-2.5"), (f"=-{bw}/2+2.5", f"=-{h}/2+2.5")]
    for i, (x, y) in enumerate(estribo):
        ws.cell(row=11 + i, column=S, value=x); ws.cell(row=11 + i, column=T, value=y)
    for base, kn, ysig in ((17, "nv", "-"), (30, "na", "+")):
        n1 = f"B{R[kn]}"
        y1c, y2c, y3c = f"B{R[kn + 'y1']}", f"B{R[kn + 'y2']}", f"B{R[kn + 'y3']}"
        for i in range(8):
            ws.cell(row=base + i, column=S,
                    value=f"=-({bw}/2-3.5)+MIN({i},{n1}-1)*2*({bw}/2-3.5)/({n1}-1)")
            ws.cell(row=base + i, column=T, value=f"={ysig}({h}/2-{y1c})")
        for j, yc in ((0, y2c), (1, y2c), (2, y3c), (3, y3c)):
            sinal = "-" if j % 2 == 0 else ""
            ws.cell(row=base + 8 + j, column=S, value=f"={sinal}({bw}/2-3.5)")
            ws.cell(row=base + 8 + j, column=T, value=f"={ysig}({h}/2-{yc})")

    ch = ScatterChart()
    ch.title = "SECAO TRANSVERSAL - vao embaixo | apoio em cima"
    ch.legend = None; ch.height = 13; ch.width = 8
    ch.x_axis.delete = True; ch.y_axis.delete = True
    ch.x_axis.majorGridlines = None; ch.y_axis.majorGridlines = None
    ch.x_axis.scaling.min = -30; ch.x_axis.scaling.max = 30
    ch.y_axis.scaling.min = -45; ch.y_axis.scaling.max = 45

    def serie(r1, r2, cor, msize=None):
        s = Series(Reference(ws, min_col=T, min_row=r1, max_row=r2),
                   Reference(ws, min_col=S, min_row=r1, max_row=r2))
        if msize:
            s.marker = Marker(symbol="circle", size=msize)
            s.marker.graphicalProperties.solidFill = cor
            s.graphicalProperties.line.noFill = True
        else:
            s.marker = Marker(symbol="none")
            s.graphicalProperties.line.solidFill = cor
        ch.series.append(s)

    serie(5, 9, "404040")
    serie(11, 15, "A6A6A6")
    serie(17, 28, "C00000", msize=9)
    serie(30, 41, "1F4E79", msize=9)
    ws.add_chart(ch, "H5")

    # ----- elevacao esquematica (apoios, armaduras, estribos, diagrama de momento) -----
    AB, AC = 28, 29
    ws.cell(row=4, column=AB, value="aux elevacao (nao apagar)").font = F_NOTA
    Lcm_, pd_, vinc_ = f"B{R['Lcm']}", f"B{R['pd']}", f"B{R['vinc']}"
    mdp_, mdn_ = f"B{R['Mdp']}", f"B{R['Mdn']}"
    sest_, y1v_, y1a_ = f"B{R['sest']}", f"B{R['nvy1']}", f"B{R['nay1']}"
    rM1, rM2, rMx, rY0 = 5, 6, 7, 8
    ws.cell(row=rM1, column=AB, value=f'=IF(OR({vinc_}="Biengastada",{vinc_}="Balanco"),-{mdn_}*100,0)')
    ws.cell(row=rM2, column=AB, value=f'=IF(OR({vinc_}="Apoiada-Engastada",{vinc_}="Biengastada"),-{mdn_}*100,0)')
    ws.cell(row=rMx, column=AB, value=f"=MAX({mdp_}*100,{mdn_}*100,1)")
    ws.cell(row=rY0, column=AB, value=f"=-({h}/2+30)")
    for i in range(40):
        ra = 10 + 2 * i
        xf = f"=IF(5+{i}*{sest_}<={Lcm_}-5,5+{i}*{sest_},NA())"
        ws.cell(row=ra, column=AB, value=xf); ws.cell(row=ra, column=AC, value=f"=-{h}/2+2.5")
        ws.cell(row=ra + 1, column=AB, value=xf); ws.cell(row=ra + 1, column=AC, value=f"={h}/2-2.5")
    for i in range(21):
        rr = 132 + i
        xi = f"({i}*{Lcm_}/20)"
        mi = (f'IF({vinc_}="Balanco",-({pd_}/100)*({Lcm_}-{xi})^2/2,'
              f'({pd_}/100)*{xi}*({Lcm_}-{xi})/2+AB{rM1}*(1-{xi}/{Lcm_})+AB{rM2}*{xi}/{Lcm_})')
        ws.cell(row=rr, column=AB, value=f"={i}*{Lcm_}/20")
        ws.cell(row=rr, column=AC, value=f"=AB{rY0}-({mi})*25/AB{rMx}")
    ws.cell(row=155, column=AB, value=0); ws.cell(row=155, column=AC, value=f"=AB{rY0}")
    ws.cell(row=156, column=AB, value=f"={Lcm_}"); ws.cell(row=156, column=AC, value=f"=AB{rY0}")
    xd1 = f'IF({vinc_}="Balanco",-15,{Lcm_})'
    xd2 = f'IF({vinc_}="Balanco",0,{Lcm_}+15)'
    apoio_esq = [("-15", f"=-{h}/2"), ("0", f"=-{h}/2"), ("0", f"=-{h}/2-20"),
                 ("-15", f"=-{h}/2-20"), ("-15", f"=-{h}/2")]
    for i, (x, y) in enumerate(apoio_esq):
        ws.cell(row=159 + i, column=AB, value=int(x)); ws.cell(row=159 + i, column=AC, value=y)
    apoio_dir = [(xd1, f"=-{h}/2"), (xd2, f"=-{h}/2"), (xd2, f"=-{h}/2-20"),
                 (xd1, f"=-{h}/2-20"), (xd1, f"=-{h}/2")]
    for i, (x, y) in enumerate(apoio_dir):
        ws.cell(row=165 + i, column=AB, value=f"={x}"); ws.cell(row=165 + i, column=AC, value=y)
    viga = [("0", f"=-{h}/2"), (f"={Lcm_}", f"=-{h}/2"), (f"={Lcm_}", f"={h}/2"),
            ("0", f"={h}/2"), ("0", f"=-{h}/2")]
    for i, (x, y) in enumerate(viga):
        v = 0 if x == "0" else x
        ws.cell(row=172 + i, column=AB, value=v); ws.cell(row=172 + i, column=AC, value=y)
    arm_pos = [("3", f"=-{h}/2+{y1v_}+8"), ("3", f"=-{h}/2+{y1v_}"),
               (f"={Lcm_}-3", f"=-{h}/2+{y1v_}"), (f"={Lcm_}-3", f"=-{h}/2+{y1v_}+8")]
    for i, (x, y) in enumerate(arm_pos):
        v = 3 if x == "3" else x
        ws.cell(row=179 + i, column=AB, value=v); ws.cell(row=179 + i, column=AC, value=y)
    arm_neg = [("3", f"={h}/2-{y1a_}-8"), ("3", f"={h}/2-{y1a_}"),
               (f"={Lcm_}-3", f"={h}/2-{y1a_}"), (f"={Lcm_}-3", f"={h}/2-{y1a_}-8")]
    for i, (x, y) in enumerate(arm_neg):
        v = 3 if x == "3" else x
        ws.cell(row=184 + i, column=AB, value=v); ws.cell(row=184 + i, column=AC, value=y)

    ch2 = ScatterChart()
    ch2.title = "ELEVACAO - armaduras, estribos, apoios e diagrama de momento (ELU)"
    ch2.legend = None; ch2.height = 11; ch2.width = 24
    ch2.x_axis.delete = True; ch2.y_axis.delete = True
    ch2.x_axis.majorGridlines = None; ch2.y_axis.majorGridlines = None

    def serie2(r1, r2, cor):
        s = Series(Reference(ws, min_col=AC, min_row=r1, max_row=r2),
                   Reference(ws, min_col=AB, min_row=r1, max_row=r2))
        s.marker = Marker(symbol="none")
        s.graphicalProperties.line.solidFill = cor
        ch2.series.append(s)

    serie2(172, 176, "008080")
    serie2(159, 163, "008080")
    serie2(165, 169, "008080")
    for i in range(40):
        serie2(10 + 2 * i, 11 + 2 * i, "00A388")
    serie2(179, 182, "C00000")
    serie2(184, 187, "1F4E79")
    serie2(155, 156, "A6A6A6")
    serie2(132, 152, "0070C0")
    ws.add_chart(ch2, "H33")
    ws.protection.sheet = True
    ws.protection.password = "dlima"
    return R


SOBRECARGAS_6120 = [
    ("Dormitorios, salas, copas, cozinhas e sanitarios (residencial)", 0.15),
    ("Despensa, area de servico e lavanderia", 0.20),
    ("Corredores dentro da unidade residencial", 0.15),
    ("Corredores de uso comum (predios)", 0.30),
    ("Escadas sem acesso ao publico", 0.25),
    ("Escadas com acesso ao publico", 0.30),
    ("Forros sem acesso a pessoas", 0.05),
    ("Coberturas acessiveis so p/ manutencao", 0.10),
    ("Escritorios (salas de uso geral)", 0.25),
    ("Salas de aula", 0.30),
    ("Lojas / comercio varejista", 0.40),
    ("Garagens (veiculos leves ate 30 kN)", 0.30),
]


def aba_laje_viga(ws):
    """Area de influencia da laje na viga (quinhao a 45 graus) + sobrecargas NBR 6120:2019."""
    titulo(ws, "LAJE -> VIGA: area de influencia + sobrecargas | D'LIMA ENGENHARIA",
           "Informe a metragem da laje; o quinhao (charneiras a 45 graus) sai sozinho | NBR 6120:2019 Tabela 10 | unidades: m, tf")
    R = {}; r = 4
    r = secao(ws, r, "LAJE 1 (apoiada nesta viga)")
    R["lx1"] = r; r = linha(ws, r, "Menor vao da laje lx", 3.0, "m", "", "in", "0.00")
    R["ly1"] = r; r = linha(ws, r, "Maior vao da laje ly", 4.0, "m", "", "in", "0.00")
    R["lado1"] = r; r = linha(ws, r, "A viga apoia qual lado da laje?", "Lado maior", "", "lado maior recebe trapezio; lado menor recebe triangulo", "in", "@")
    R["A1"] = r; r = linha(ws, r, "Quinhao da laje 1", f'=IF(B{R["lado1"]}="Lado maior",MIN(B{R["lx1"]},B{R["ly1"]})*(2*MAX(B{R["lx1"]},B{R["ly1"]})-MIN(B{R["lx1"]},B{R["ly1"]}))/4,MIN(B{R["lx1"]},B{R["ly1"]})^2/4)', "m2", "trapezio: lx*(2*ly-lx)/4 | triangulo: lx^2/4")
    r += 1
    r = secao(ws, r, "LAJE 2 (do outro lado da viga, se houver)")
    R["tem2"] = r; r = linha(ws, r, "Ha laje do outro lado?", "Nao", "", "Sim / Nao", "in", "@")
    R["lx2"] = r; r = linha(ws, r, "Menor vao da laje lx", 3.0, "m", "", "in", "0.00")
    R["ly2"] = r; r = linha(ws, r, "Maior vao da laje ly", 4.0, "m", "", "in", "0.00")
    R["lado2"] = r; r = linha(ws, r, "A viga apoia qual lado da laje?", "Lado maior", "", "", "in", "@")
    R["A2"] = r; r = linha(ws, r, "Quinhao da laje 2", f'=IF(B{R["tem2"]}="Nao",0,IF(B{R["lado2"]}="Lado maior",MIN(B{R["lx2"]},B{R["ly2"]})*(2*MAX(B{R["lx2"]},B{R["ly2"]})-MIN(B{R["lx2"]},B{R["ly2"]}))/4,MIN(B{R["lx2"]},B{R["ly2"]})^2/4))', "m2")
    r += 1
    R["At"] = r
    ws.cell(row=r, column=1, value="AREA DE INFLUENCIA TOTAL -> copie p/ a aba DIM VIGA").font = Font(bold=True, size=12)
    c = ws.cell(row=r, column=2, value=f"=B{R['A1']}+B{R['A2']}")
    c.font = Font(bold=True, size=12); c.border = borda; c.number_format = "0.00"
    ws.cell(row=r, column=3, value="m2").font = F_NOTA
    r += 2
    dv_lista(ws, f"B{R['lado1']}", "Lado maior,Lado menor")
    dv_lista(ws, f"B{R['lado2']}", "Lado maior,Lado menor")
    dv_lista(ws, f"B{R['tem2']}", "Sim,Nao")
    ws.column_dimensions["E"].width = 36
    ws.cell(row=3, column=5, value="NORMA P/ CONSULTA RAPIDA").font = Font(bold=True, size=9, color="1A3C6E")
    for k in ("lx1", "ly1", "lx2", "ly2"):
        ws.cell(row=R[k], column=5, value="NBR 6118 item 14.7.2.2 (vao efetivo da laje)").font = F_NOTA
    for k in ("lado1", "lado2"):
        ws.cell(row=R[k], column=5, value="charneiras a 45 graus (Pinheiro, Fund. do Concreto, cap. lajes)").font = F_NOTA

    r = secao(ws, r, "SOBRECARGA DE USO - NBR 6120:2019 (Tabela 10)")
    for j, t in enumerate(("Uso do ambiente sobre a laje", "q (tf/m2)", "q (kN/m2)"), start=1):
        c = ws.cell(row=r, column=j, value=t)
        c.font = Font(bold=True, size=9); c.fill = PatternFill("solid", start_color=CINZA); c.border = borda
    r += 1
    for uso, q in SOBRECARGAS_6120:
        ws.cell(row=r, column=1, value=uso).border = borda
        c = ws.cell(row=r, column=2, value=q); c.border = borda; c.number_format = "0.00"
        c2 = ws.cell(row=r, column=3, value=f"=B{r}*10"); c2.border = borda; c2.number_format = "0.0"
        r += 1
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1, value="Sacadas e varandas: usar a mesma carga do ambiente que se comunica (NBR 6120). Use o valor em tf/m2 no campo 'Carga acidental da laje' da aba DIM VIGA.").font = F_NOTA
    r += 2
    notas = [
        "Quinhao a 45 graus (metodo dos triangulos e trapezios, todas as bordas apoiadas): e o metodo usual de descarga de laje macica em viga.",
        "Se a laje for muito alongada (ly/lx > 2), a laje e armada em 1 direcao: as vigas do lado MAIOR recebem quase toda a carga (o quinhao do lado menor e pequeno e este metodo continua valido).",
        "Vigas intermediarias recebem laje dos DOIS lados: preencha a Laje 2.",
    ]
    for n in notas:
        ws.merge_cells(f"A{r}:E{r}")
        ws.cell(row=r, column=1, value=n).font = Font(size=9)
        r += 1
    ws.protection.sheet = True
    ws.protection.password = "dlima"
    return R


def main():
    wb = Workbook()
    ws_dash = wb.active; ws_dash.title = "DASHBOARD"
    ws_rap = wb.create_sheet("RAPIDO")
    ws_vig = wb.create_sheet("VIGA")
    ws_dim = wb.create_sheet("DIM VIGA")
    ws_lv = wb.create_sheet("LAJE-VIGA")
    ws_lm = wb.create_sheet("LAJE MACICA")
    ws_lt = wb.create_sheet("LAJE TRELICADA")
    ws_nor = wb.create_sheet("NORMA")

    R_rap = aba_rapido(ws_rap)
    R_vig = aba_retangular(ws_vig, "VIGA DE CONCRETO ARMADO", dict(
        L=4.0, vinc="Biapoiada", fck=25, agreg="Granito/Gnaisse", b=14.0, h=40.0, d=36.0, dl=3.0,
        g=1.2, q=0.4, psi2=0.3, As=3.68, Asl=1.01, t0=1.0, parede="Sim",
        as_nota=" no vao", as_nota2="ex.: 3 fi 12,5 = 3,68"), "tf/m", "carga por metro de viga")
    aba_dimensionamento_viga(ws_dim)
    aba_laje_viga(ws_lv)
    R_lm = aba_retangular(ws_lm, "LAJE MACICA (faixa de 1,00 m)", dict(
        L=4.0, vinc="Biapoiada", fck=25, agreg="Granito/Gnaisse", h=12.0, d=9.5, dl=2.5,
        g=0.45, q=0.15, psi2=0.3, As=3.35, Asl=0.0, t0=1.0, parede="Nao",
        as_nota=" por metro", as_nota2="ex.: fi 8 c/15 = 3,35 cm2/m"), "tf/m2", "por m2 (a faixa de 1 m converte sozinha)", b_fixo=100)
    R_lt = aba_trelicada(ws_lt)
    aba_dashboard(ws_dash, [("MODO RAPIDO (flecha do software)", "RAPIDO", R_rap),
                            ("VIGA", "VIGA", R_vig),
                            ("LAJE MACICA", "LAJE MACICA", R_lm),
                            ("LAJE TRELICADA", "LAJE TRELICADA", R_lt)])
    aba_norma(ws_nor)

    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = "dlima"

    saida = r"C:\Users\leona\Desktop\DLIMA\Flechas_NBR6118_DLIMA.xlsx"
    wb.save(saida)
    print("OK:", saida)


if __name__ == "__main__":
    main()
