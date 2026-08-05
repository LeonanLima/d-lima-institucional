# -*- coding: utf-8 -*-
"""Adiciona a aba DIM VIGA (dimensionamento ELU) na planilha de flechas existente."""
import sys
from openpyxl import load_workbook

from gerar_planilha_flecha import aba_dimensionamento_viga, aba_laje_viga

ARQ = r"C:\Users\leona\Desktop\DLIMA\Flechas_NBR6118_DLIMA.xlsx"


def main() -> None:
    wb = load_workbook(ARQ)
    for nome in ("DIM VIGA", "LAJE-VIGA"):
        if nome in wb.sheetnames:
            del wb[nome]
    pos = wb.sheetnames.index("VIGA") + 1 if "VIGA" in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet("DIM VIGA", pos)
    aba_dimensionamento_viga(ws)
    aba_laje_viga(wb.create_sheet("LAJE-VIGA", pos + 1))
    try:
        wb.save(ARQ)
    except PermissionError:
        sys.exit("ERRO: feche a planilha no Excel e rode de novo.")
    print("OK:", ARQ)


if __name__ == "__main__":
    main()
